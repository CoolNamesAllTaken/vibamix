"""Parse an XLSX bulk-programming sheet into per-badge content specs.

One row per badge (applied in table order to the checked devices). Columns are
identified by *named* headers (case-insensitive, order-independent); unoccupied
frames are simply omitted. See README.md "Bulk (XLSX) programming" for the schema.

Recognized headers:
  name                       identity display name (<=31 B)
  attendee_id|attendee|table_id  identity attendee/table id (<=10 B)
  identity_image             path to the identity background image
  identity_image_fmt         bw|gray2 (default gray2)
  text_label_N / text_body_N text frame N (N=1..20 -> slot idx N-1)
  image_N                    path for image slot N (N=1..4 -> slot N-1)
  image_N_fmt                bw|gray2 for slot N (default gray2)

This module is pure (no Qt / BLE) so it can be unit-tested in isolation.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from . import keys

MAX_TEXT_FRAMES = 20   # firmware: 20 text screens (idx 0..19)
MAX_IMAGE_SLOTS = 4    # firmware: 4 user image slots (slot 0..3)
NAME_MAX = 31
ATTENDEE_MAX = 10

_RE_TEXT_LABEL = re.compile(r"^text_label_(\d+)$")
_RE_TEXT_BODY = re.compile(r"^text_body_(\d+)$")
_RE_TEXT_LED = re.compile(r"^text_led_(\d+)$")
_RE_TEXT_ANIM = re.compile(r"^text_anim_(\d+)$")
_RE_IMAGE_FMT = re.compile(r"^image_(\d+)_fmt$")
_RE_IMAGE_LED = re.compile(r"^image_led_(\d+)$")
_RE_IMAGE_ANIM = re.compile(r"^image_anim_(\d+)$")
_RE_IMAGE = re.compile(r"^image_(\d+)$")

_ATTENDEE_ALIASES = {"attendee_id", "attendee", "table_id"}

# CSS-ish color names; everything else falls back to hex (#rrggbb / rrggbb / rgb).
_COLOR_NAMES = {
    "black": (0, 0, 0), "white": (255, 255, 255), "red": (255, 0, 0),
    "green": (0, 255, 0), "lime": (0, 255, 0), "blue": (0, 0, 255),
    "yellow": (255, 255, 0), "cyan": (0, 255, 255), "aqua": (0, 255, 255),
    "magenta": (255, 0, 255), "fuchsia": (255, 0, 255), "orange": (255, 165, 0),
    "purple": (128, 0, 128), "pink": (255, 192, 203), "teal": (0, 128, 128),
    "navy": (0, 0, 128), "maroon": (128, 0, 0), "olive": (128, 128, 0),
    "gray": (128, 128, 128), "grey": (128, 128, 128), "silver": (192, 192, 192),
    "gold": (255, 215, 0), "violet": (238, 130, 238), "indigo": (75, 0, 130),
}
# animation name (spaces/underscores stripped) -> code, from keys.ANIM_NAMES.
_ANIM_BY_NAME = {name.lower().replace(" ", "").replace("_", ""): code
                 for code, name in keys.ANIM_NAMES.items()}


@dataclass
class TextFrameSpec:
    idx: int
    label: str
    body: str
    anim: int = keys.ANIM_OFF
    r: int = 0
    g: int = 0
    b: int = 0


@dataclass
class ImageFrameSpec:
    slot: int
    path: str            # absolute, resolved against the workbook directory
    fmt: int             # keys.FMT_BW | keys.FMT_GRAY2
    anim: int = keys.ANIM_OFF
    r: int = 0
    g: int = 0
    b: int = 0


@dataclass
class RowSpec:
    row_no: int          # 1-based spreadsheet row (for messages)
    name: str | None = None
    attendee_id: str | None = None
    identity_image: tuple[str, int] | None = None   # (abs path, fmt)
    identity_led: tuple[int, int, int, int] | None = None   # (anim, r, g, b)
    text_frames: list[TextFrameSpec] = field(default_factory=list)
    image_frames: list[ImageFrameSpec] = field(default_factory=list)

    def is_empty(self) -> bool:
        return (self.name is None and self.attendee_id is None
                and self.identity_image is None and self.identity_led is None
                and not self.text_frames and not self.image_frames)


def _norm(header: object) -> str:
    return str(header).strip().lower().replace(" ", "_") if header is not None else ""


def _cell_str(v: object) -> str:
    """Normalize a cell value to a stripped string ('' for blank/None)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _fmt_from_str(s: str, default: int, warnings: list[str], where: str) -> int:
    t = s.strip().lower()
    if t in ("", "gray2", "grey2", "gray", "grey", "g2"):
        return keys.FMT_GRAY2 if t else default
    if t in ("bw", "b&w", "1bpp", "mono"):
        return keys.FMT_BW
    warnings.append(f"{where}: unknown image format '{s}', using "
                    f"{'gray2' if default == keys.FMT_GRAY2 else 'bw'}")
    return default


def _parse_color(s: str, warnings: list[str], where: str) -> tuple[int, int, int] | None:
    """A named color or hex (#rrggbb / rrggbb / rgb). Blank -> None; bad -> warn + None."""
    t = s.strip().lower()
    if not t:
        return None
    if t in _COLOR_NAMES:
        return _COLOR_NAMES[t]
    h = t[1:] if t.startswith("#") else t
    if len(h) == 3 and all(c in "0123456789abcdef" for c in h):
        h = "".join(c * 2 for c in h)
    if len(h) == 6 and all(c in "0123456789abcdef" for c in h):
        return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))
    warnings.append(f"{where}: unrecognized color '{s}' (ignored)")
    return None


def _parse_anim(s: str, warnings: list[str], where: str) -> int | None:
    """An animation name (spaces/underscores ignored) or its 0..7 code."""
    t = s.strip().lower()
    if not t:
        return None
    key = t.replace(" ", "").replace("_", "")
    if key in _ANIM_BY_NAME:
        return _ANIM_BY_NAME[key]
    if key.isdigit() and int(key) in keys.ANIM_NAMES:
        return int(key)
    warnings.append(f"{where}: unknown animation '{s}' (ignored)")
    return None


def _resolve_led(color_s: str, anim_s: str, warnings: list[str],
                 where: str) -> tuple[int, int, int, int] | None:
    """Combine an LED color + animation cell into (anim, r, g, b).

    color but no anim -> Solid; anim but no color -> that anim on black (warn if
    Solid, which is just off); neither -> None (caller keeps the OFF default).
    """
    color = _parse_color(color_s, warnings, where)
    anim = _parse_anim(anim_s, warnings, where)
    if color is None and anim is None:
        return None
    if anim is None:
        anim = keys.ANIM_SOLID
    if color is None:
        if anim == keys.ANIM_SOLID:
            warnings.append(f"{where}: solid animation with no color shows as off")
        color = (0, 0, 0)
    return (anim, *color)


def _resolve(path: str, base: Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else (base / p)


def _image_load_error(p: Path) -> str | None:
    """Return an error string if the image is missing or can't be decoded, else None."""
    if not p.is_file():
        return "not found"
    try:
        from PIL import Image  # local import: keep the parser import light
        with Image.open(p) as im:
            im.load()
    except Exception as e:  # noqa: BLE001
        return f"cannot load image ({e})"
    return None


def parse_workbook(xlsx_path: str | Path) -> tuple[list[RowSpec], list[str]]:
    """Parse the active sheet into (rows, warnings).

    Image cells are resolved against the workbook's directory and each one is
    validated by actually opening it. If ANY referenced image is missing or can't be
    decoded, the whole parse fails with a ValueError listing every bad image (so the
    operator fixes them before starting a run with badges connected). Wholly-empty
    rows are skipped.
    """
    import openpyxl  # local import: keep the dependency lazy

    xlsx_path = Path(xlsx_path)
    base = xlsx_path.resolve().parent
    warnings: list[str] = []
    errors: list[str] = []

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            return [], ["sheet is empty"]

        cols: dict[str, int] = {}
        text_label: dict[int, int] = {}   # frame N -> column index
        text_body: dict[int, int] = {}
        text_led: dict[int, int] = {}
        text_anim: dict[int, int] = {}
        image_path: dict[int, int] = {}   # slot N -> column index
        image_fmt: dict[int, int] = {}
        image_led: dict[int, int] = {}
        image_anim: dict[int, int] = {}
        for ci, raw in enumerate(header):
            h = _norm(raw)
            if not h:
                continue
            if h in _ATTENDEE_ALIASES:
                cols["attendee_id"] = ci
            elif h in ("name", "identity_image", "identity_image_fmt",
                       "identity_led", "identity_anim"):
                cols[h] = ci
            elif (m := _RE_TEXT_LABEL.match(h)):
                text_label[int(m.group(1))] = ci
            elif (m := _RE_TEXT_BODY.match(h)):
                text_body[int(m.group(1))] = ci
            elif (m := _RE_TEXT_LED.match(h)):
                text_led[int(m.group(1))] = ci
            elif (m := _RE_TEXT_ANIM.match(h)):
                text_anim[int(m.group(1))] = ci
            elif (m := _RE_IMAGE_FMT.match(h)):
                image_fmt[int(m.group(1))] = ci
            elif (m := _RE_IMAGE_LED.match(h)):
                image_led[int(m.group(1))] = ci
            elif (m := _RE_IMAGE_ANIM.match(h)):
                image_anim[int(m.group(1))] = ci
            elif (m := _RE_IMAGE.match(h)):
                image_path[int(m.group(1))] = ci
            else:
                warnings.append(f"ignoring unrecognized column '{raw}'")

        def cell(row: tuple, ci: int | None) -> str:
            if ci is None or ci >= len(row):
                return ""
            return _cell_str(row[ci])

        rows: list[RowSpec] = []
        for ri, row in enumerate(it, start=2):  # data starts at spreadsheet row 2
            spec = RowSpec(row_no=ri)
            name = cell(row, cols.get("name"))
            att = cell(row, cols.get("attendee_id"))
            if name:
                if len(name.encode()) > NAME_MAX:
                    warnings.append(f"row {ri}: name longer than {NAME_MAX} B, will be truncated")
                spec.name = name
            if att:
                if len(att.encode()) > ATTENDEE_MAX:
                    warnings.append(f"row {ri}: attendee_id longer than {ATTENDEE_MAX} B, "
                                    "will be truncated")
                spec.attendee_id = att

            id_img = cell(row, cols.get("identity_image"))
            if id_img:
                fmt = _fmt_from_str(cell(row, cols.get("identity_image_fmt")),
                                    keys.FMT_GRAY2, warnings, f"row {ri} identity_image")
                p = _resolve(id_img, base)
                err = _image_load_error(p)
                if err:
                    errors.append(f"row {ri}: identity_image: {p}: {err}")
                else:
                    spec.identity_image = (str(p), fmt)

            id_led = _resolve_led(cell(row, cols.get("identity_led")),
                                  cell(row, cols.get("identity_anim")),
                                  warnings, f"row {ri} identity")
            if id_led is not None:
                if spec.name is None and spec.attendee_id is None:
                    warnings.append(f"row {ri}: identity LED ignored "
                                    "(set name or attendee_id in the same row)")
                else:
                    spec.identity_led = id_led

            for n in sorted(set(text_label) | set(text_body) | set(text_led) | set(text_anim)):
                if not (1 <= n <= MAX_TEXT_FRAMES):
                    warnings.append(f"text frame {n} out of range 1..{MAX_TEXT_FRAMES} (ignored)")
                    continue
                label = cell(row, text_label.get(n))
                body = cell(row, text_body.get(n))
                led = _resolve_led(cell(row, text_led.get(n)), cell(row, text_anim.get(n)),
                                   warnings, f"row {ri} text_{n}")
                if label or body:
                    a, r, g, b = led or (keys.ANIM_OFF, 0, 0, 0)
                    spec.text_frames.append(
                        TextFrameSpec(idx=n - 1, label=label, body=body, anim=a, r=r, g=g, b=b))
                elif led is not None:
                    warnings.append(f"row {ri}: text_{n} LED ignored (no label/body)")

            for n in sorted(set(image_path) | set(image_led) | set(image_anim)):
                if not (1 <= n <= MAX_IMAGE_SLOTS):
                    warnings.append(f"image slot {n} out of range 1..{MAX_IMAGE_SLOTS} (ignored)")
                    continue
                path = cell(row, image_path.get(n))
                led = _resolve_led(cell(row, image_led.get(n)), cell(row, image_anim.get(n)),
                                   warnings, f"row {ri} image_{n}")
                if not path:
                    if led is not None:
                        warnings.append(f"row {ri}: image_{n} LED ignored (no image path)")
                    continue
                fmt = _fmt_from_str(cell(row, image_fmt.get(n)),
                                    keys.FMT_GRAY2, warnings, f"row {ri} image_{n}")
                p = _resolve(path, base)
                err = _image_load_error(p)
                if err:
                    errors.append(f"row {ri}: image_{n}: {p}: {err}")
                    continue
                a, r, g, b = led or (keys.ANIM_OFF, 0, 0, 0)
                spec.image_frames.append(
                    ImageFrameSpec(slot=n - 1, path=str(p), fmt=fmt, anim=a, r=r, g=g, b=b))

            if not spec.is_empty():
                rows.append(spec)
    finally:
        wb.close()

    if errors:
        raise ValueError("could not load "
                         + ("1 image:\n  " if len(errors) == 1
                            else f"{len(errors)} images:\n  ")
                         + "\n  ".join(errors))
    return rows, warnings


# --- header layout for a starter template -----------------------------------

TEMPLATE_HEADERS = [
    "name", "attendee_id", "identity_image", "identity_image_fmt",
    "identity_led", "identity_anim",
    "text_label_1", "text_body_1", "text_led_1", "text_anim_1",
    "text_label_2", "text_body_2",
    "image_1", "image_1_fmt", "image_led_1", "image_anim_1",
    "image_2", "image_2_fmt",
]


def make_template(path: str | Path) -> None:
    """Write a starter workbook: the recognized header row + one example row."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "badges"
    ws.append(TEMPLATE_HEADERS)
    # Example row. Image-path cells are left blank so a freshly-saved template loads
    # without error (image paths are validated on load); fill them with real paths.
    ws.append([
        "Ada Lovelace", "T12", "", "",
        "teal", "solid",
        "About me", "First programmer.", "orange", "wheel",
        "Ask me about", "Analytical engines.",
        "", "", "", "",
        "", "",
    ])
    wb.save(str(path))
